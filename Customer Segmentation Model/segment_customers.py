import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(
    title="Select Customer Excel File",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

if not file_path:
    messagebox.showwarning("No File Selected", "You must select an Excel file.")
    exit()

df = pd.read_excel(file_path)
df.columns = df.columns.str.strip()

print("Available columns:", df.columns.tolist())

income_col = next((col for col in df.columns if '$' in col.lower()), None)
spending_col = next((col for col in df.columns if 'spending' in col.lower()), None)
age_col = next((col for col in df.columns if 'age' in col.lower()), None)

if not all([income_col, spending_col, age_col]):
    messagebox.showerror("Missing Columns", "The file must include Age, Income ($), and Spending Score columns.")
    exit()

features = [age_col, income_col, spending_col]
data = df[features]

scaler = StandardScaler()
scaled_data = scaler.fit_transform(data)

kmeans = KMeans(n_clusters=4, random_state=42)
clusters = kmeans.fit_predict(scaled_data)

df['Cluster'] = clusters

def interpret_cluster(row):
    if row['Cluster'] == 0:
        return 'High Income, High Spending'
    elif row['Cluster'] == 1:
        return 'Low Income, Low Spending'
    elif row['Cluster'] == 2:
        return 'High Income, Low Spending'
    else:
        return 'Young Spenders'

df['Cluster Profile'] = df.apply(interpret_cluster, axis=1)

plt.figure(figsize=(10, 6))
plt.scatter(df[income_col], df[spending_col], c=df['Cluster'], cmap='viridis', s=50)
plt.xlabel(income_col)
plt.ylabel(spending_col)
plt.title('Customer Segmentation')
plt.grid(True)
plt.tight_layout()
plt.show()

export_path = os.path.splitext(file_path)[0] + "_segmented.xlsx"
df.to_excel(export_path, index=False)

wb = load_workbook(export_path)
ws = wb.active

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
cell_fill_even = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
cell_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
center_alignment = Alignment(horizontal='center')

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        cell.alignment = center_alignment
        cell.border = thin_border
        
        if cell.row == 1:
            cell.font = header_font
            cell.fill = header_fill
        else:
            if cell.row % 2 == 0:
                cell.fill = cell_fill_even
            else:
                cell.fill = cell_fill_odd
        
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            if cell.column not in [ws.max_column, ws.max_column-1, 2]:
                cell.number_format = '#,##0'

ws.freeze_panes = 'A2'

wb.save(export_path)

print("Export complete with enhanced formatting!")
print(df.head())